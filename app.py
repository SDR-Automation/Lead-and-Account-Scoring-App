import streamlit as st
import pandas as pd
import json
import re
import os
from groq import Groq
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
client = Groq(api_key=GROQ_API_KEY)

st.set_page_config(
    page_title="Ebryx Lead Intelligence",
    page_icon="●",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');

:root {
    --black:        #0E0E0E;
    --black-2:      #141414;
    --black-3:      #1a1a1a;
    --black-4:      #222222;
    --black-5:      #2d2d2d;
    --red:          #C41A1A;
    --red-bright:   #FF4040;
    --red-light:    rgba(196,26,26,0.12);
    --red-border:   rgba(196,26,26,0.3);
    --white:        #FFFFFF;
    --grey-1:       #F0F0F0;
    --grey-2:       #A0A0A0;
    --grey-3:       #5a5a5a;
    --grey-4:       #3a3a3a;
    --border:       #2a2a2a;
    --border-red:   rgba(196,26,26,0.4);
}

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
    background-color: var(--black);
    color: var(--grey-1);
}

.stApp { background: var(--black); }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 !important; max-width: 100% !important; }

/* ── TOP NAV ── */
.topnav {
    background: var(--black-2);
    border-bottom: 1px solid var(--border);
    padding: 0 48px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    height: 60px;
}
.topnav-left { display: flex; align-items: center; gap: 32px; }
.topnav-brand {
    font-family: 'Bebas Neue', sans-serif;
    font-size: 1.4rem; letter-spacing: 0.12em; color: var(--white);
    display: flex; align-items: center; gap: 10px;
}
.topnav-dot {
    width: 10px; height: 10px; background: var(--red);
    border-radius: 50%; display: inline-block;
    box-shadow: 0 0 8px rgba(196,26,26,0.6);
}
.topnav-divider { width: 1px; height: 24px; background: var(--border); }
.topnav-tag {
    font-family: 'DM Mono', monospace; font-size: 0.68rem;
    color: var(--grey-3); letter-spacing: 0.1em; text-transform: uppercase;
    border: 1px solid var(--border); padding: 4px 10px; border-radius: 4px;
}
.topnav-right {
    display: flex; align-items: center; gap: 8px;
    font-family: 'DM Mono', monospace; font-size: 0.68rem;
    color: var(--red-bright); letter-spacing: 0.1em;
}
.live-dot {
    width: 7px; height: 7px; background: var(--red-bright);
    border-radius: 50%; animation: blink 1.5s infinite;
}
@keyframes blink { 0%,100%{opacity:1} 50%{opacity:0.3} }

/* ── HERO ── */
.hero {
    background: var(--black);
    padding: 60px 48px 48px;
    border-bottom: 1px solid var(--border);
    position: relative; overflow: hidden;
    text-align: center;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
}
.hero::after {
    content: ''; position: absolute; bottom: 0; left: 0; right: 0; height: 1px;
    background: linear-gradient(90deg, transparent, var(--red), transparent);
}
.hero-label {
    font-family: 'DM Mono', monospace; font-size: 0.68rem; color: var(--red-bright);
    letter-spacing: 0.2em; text-transform: uppercase; margin-bottom: 16px;
    display: flex; align-items: center; justify-content: center; gap: 10px;
}
.hero-label::before { content: ''; width: 20px; height: 1px; background: var(--red); }
.hero-label::after  { content: ''; width: 20px; height: 1px; background: var(--red); }
.hero-h1 {
    font-family: 'Bebas Neue', sans-serif; font-size: 5rem; line-height: 0.95;
    letter-spacing: 0.03em; color: var(--white); margin-bottom: 20px;
    text-align: center;
}
.hero-h1 .red { color: var(--red-bright); }
.hero-p {
    font-size: 0.95rem; color: var(--grey-2); max-width: 500px;
    line-height: 1.8; font-weight: 300; text-align: center; margin: 0 auto;
}
.hero-bg-text {
    position: absolute; left: 50%; top: 50%;
    transform: translate(-50%, -50%);
    font-family: 'Bebas Neue', sans-serif; font-size: 14rem;
    color: rgba(196,26,26,0.03); letter-spacing: 0.05em;
    user-select: none; pointer-events: none; line-height: 1; white-space: nowrap;
}

/* ── SECTION HEADERS ── */
.sec-hdr {
    font-family: 'DM Mono', monospace; font-size: 0.62rem; font-weight: 500;
    color: var(--grey-3); letter-spacing: 0.2em; text-transform: uppercase;
    margin-bottom: 20px; display: flex; align-items: center; gap: 12px;
}
.sec-hdr::after { content: ''; flex: 1; height: 1px; background: var(--border); }

/* ── INFO CARD ── */
.req-card {
    background: var(--black-3); border: 1px solid var(--border);
    border-top: 2px solid var(--red); border-radius: 6px;
    padding: 16px 18px; margin-bottom: 20px;
}
.req-card-title {
    font-family: 'DM Mono', monospace; font-size: 0.62rem; color: var(--grey-3);
    letter-spacing: 0.15em; text-transform: uppercase; margin-bottom: 12px;
}
.col-pill {
    display: inline-block; background: var(--red-light);
    border: 1px solid var(--red-border); color: var(--red-bright);
    padding: 3px 10px; border-radius: 3px;
    font-family: 'DM Mono', monospace; font-size: 0.72rem; margin: 2px; letter-spacing: 0.05em;
}

/* ── METRICS ── */
.metrics-row {
    display: grid; grid-template-columns: repeat(5, 1fr);
    gap: 10px; margin-bottom: 24px;
}
.met-card {
    background: var(--black-2); border: 1px solid var(--border);
    border-radius: 6px; padding: 16px 12px; text-align: center; position: relative;
}
.met-card::after {
    content: ''; position: absolute; bottom: 0; left: 0; right: 0;
    height: 2px; border-radius: 0 0 6px 6px;
}
.met-card.c-total::after { background: var(--border); }
.met-card.c-vh::after    { background: var(--red-bright); }
.met-card.c-high::after  { background: #FF8C00; }
.met-card.c-med::after   { background: #A0A0A0; }
.met-card.c-low::after   { background: var(--black-5); }
.met-num { font-family: 'Bebas Neue', sans-serif; font-size: 2.2rem; line-height: 1; margin-bottom: 2px; }
.met-lbl { font-family: 'DM Mono', monospace; font-size: 0.58rem; color: var(--grey-3); letter-spacing: 0.15em; text-transform: uppercase; }

/* ── SCORE BARS ── */
.sb-row { display: flex; align-items: center; gap: 10px; margin-bottom: 9px; }
.sb-lbl { font-family: 'DM Mono', monospace; font-size: 0.65rem; color: var(--grey-3); width: 140px; flex-shrink: 0; letter-spacing: 0.04em; }
.sb-track { flex: 1; height: 4px; background: var(--black-5); border-radius: 2px; overflow: hidden; }
.sb-fill  { height: 100%; border-radius: 2px; }
.sb-val   { font-family: 'Bebas Neue', sans-serif; font-size: 1rem; width: 22px; text-align: right; letter-spacing: 0.05em; }

/* ── DETAIL CARDS ── */
.detail-label { font-family: 'DM Mono', monospace; font-size: 0.6rem; color: var(--grey-3); text-transform: uppercase; letter-spacing: 0.15em; margin-bottom: 5px; }
.detail-value { font-size: 0.88rem; color: var(--grey-1); font-weight: 500; }
.rationale-box {
    background: var(--black-3); border: 1px solid var(--border);
    border-left: 2px solid var(--grey-4); border-radius: 4px;
    padding: 14px 16px; font-size: 0.8rem; color: var(--grey-2); line-height: 1.7;
}
.action-box {
    background: var(--red-light); border: 1px solid var(--red-border);
    border-left: 2px solid var(--red); border-radius: 4px;
    padding: 14px 16px; font-size: 0.8rem; color: var(--grey-1); line-height: 1.7; font-weight: 500;
}

/* ── UPLOAD PLACEHOLDER ── */
.up-placeholder { border: 1px dashed var(--border); border-radius: 6px; padding: 32px 20px; text-align: center; }

/* ── EMPTY STATE ── */
.empty-st { display: flex; flex-direction: column; align-items: center; justify-content: center; padding: 100px 40px; text-align: center; }
.empty-icon { font-family: 'Bebas Neue', sans-serif; font-size: 5rem; color: var(--black-4); line-height: 1; margin-bottom: 16px; }
.empty-title { font-family: 'Bebas Neue', sans-serif; font-size: 1.6rem; color: var(--black-5); letter-spacing: 0.08em; margin-bottom: 8px; }
.empty-sub { font-size: 0.8rem; color: var(--grey-3); line-height: 1.8; max-width: 280px; }

/* ── STREAMLIT OVERRIDES ── */
.stFileUploader > div { background: var(--black-3) !important; border: 1px dashed var(--border) !important; border-radius: 6px !important; }
.stFileUploader label { color: var(--grey-2) !important; font-family: 'DM Mono', monospace !important; font-size: 0.68rem !important; letter-spacing: 0.15em !important; text-transform: uppercase !important; }

.stButton > button {
    background: var(--red) !important; color: var(--white) !important;
    font-family: 'Bebas Neue', sans-serif !important; font-size: 1rem !important;
    letter-spacing: 0.15em !important; border: none !important; border-radius: 4px !important;
    padding: 14px 24px !important; width: 100% !important; text-transform: uppercase !important;
    transition: all 0.15s ease !important;
}
.stButton > button:hover { background: var(--red-bright) !important; box-shadow: 0 4px 20px rgba(196,26,26,0.4) !important; transform: translateY(-1px) !important; }

.stDownloadButton > button {
    background: var(--black-3) !important; color: var(--grey-1) !important;
    font-family: 'Bebas Neue', sans-serif !important; font-size: 1rem !important;
    letter-spacing: 0.15em !important; border: 1px solid var(--border-red) !important;
    border-radius: 4px !important; padding: 14px 24px !important; width: 100% !important;
    text-transform: uppercase !important; transition: all 0.15s ease !important;
}
.stDownloadButton > button:hover { background: var(--red-light) !important; border-color: var(--red) !important; box-shadow: 0 4px 20px rgba(196,26,26,0.2) !important; }

.stProgress > div > div { background: linear-gradient(90deg, var(--red), var(--red-bright)) !important; border-radius: 2px !important; }
.stProgress > div { background: var(--black-4) !important; border-radius: 2px !important; height: 4px !important; }

div[data-testid="stDataFrame"] { border-radius: 6px !important; border: 1px solid var(--border) !important; overflow: hidden !important; }

.stSelectbox > div > div { background: var(--black-3) !important; border: 1px solid var(--border) !important; border-radius: 4px !important; color: var(--grey-1) !important; font-family: 'DM Mono', monospace !important; font-size: 0.78rem !important; }
.stMultiSelect > div > div { background: var(--black-3) !important; border: 1px solid var(--border) !important; border-radius: 4px !important; }

div[data-testid="stExpander"] { background: var(--black-2) !important; border: 1px solid var(--border) !important; border-radius: 4px !important; margin-bottom: 6px !important; }
div[data-testid="stExpander"]:hover { border-color: var(--border-red) !important; }

label[data-testid="stWidgetLabel"] { color: var(--grey-3) !important; font-family: 'DM Mono', monospace !important; font-size: 0.62rem !important; font-weight: 500 !important; letter-spacing: 0.15em !important; text-transform: uppercase !important; }

.stSuccess > div { background: rgba(196,26,26,0.06) !important; border: 1px solid rgba(196,26,26,0.2) !important; border-radius: 4px !important; color: var(--red-bright) !important; font-family: 'DM Mono', monospace !important; font-size: 0.78rem !important; }
.stWarning > div { background: rgba(255,140,0,0.06) !important; border: 1px solid rgba(255,140,0,0.2) !important; border-radius: 4px !important; }
.stError > div   { background: rgba(196,26,26,0.08) !important; border: 1px solid rgba(196,26,26,0.3) !important; border-radius: 4px !important; }
</style>
""", unsafe_allow_html=True)


def score_color_ui(v):
    if v >= 8:   return "#FF4040"
    elif v >= 6: return "#FF8C00"
    elif v >= 4: return "#A0A0A0"
    else:        return "#5a5a5a"


def build_prompt(companies):
    companies_text = "\n".join(
        f"{i+1}. {json.dumps(c, ensure_ascii=False)}" for i, c in enumerate(companies)
    )
    return f"""
You are an expert B2B sales intelligence analyst for Ebryx, a cybersecurity firm specializing in AppSec and compliance testing.
For each company below, you are given only the company name and website.
You must use your own knowledge to research each company and infer all details.

COMPANIES TO ANALYZE:
{companies_text}

Research and provide: specialities, industry, estimated revenue, employee size, region, funding stage.

Score each on 5 dimensions (0-10):
1. STRATEGIC FIT: 0-2=Outside ICP, 3-5=Partial, 6-8=Strong, 9-10=Ideal ICP
2. REVENUE POTENTIAL: 0-2=<50emp, 3-4=50-100, 5-6=100-300, 7-8=300-1000, 9-10=Enterprise
3. REGULATORY PRESSURE: 0-2=Unregulated, 3-4=Mild, 5-6=SOC2/ISO, 7-8=PCI/GDPR, 9-10=Active audits
4. CYBERSECURITY CRITICALITY: 0-2=Internal, 3-4=Limited, 5-6=Customer-facing, 7-8=Financial, 9-10=Mission-critical
5. VALUE-ADD: 0-2=Commodity, 3-4=Some diff, 5-6=AppSec overlap, 7-8=Clear edge, 9-10=Unique

INDEX SCORE = sum of all 5 (max 50). 40-50=Very High, 30-39=High, 20-29=Medium, 0-19=Low

Return ONLY a valid JSON array. No markdown. Start with [ end with ]
Each object must have: company_name, website, specialities, industry, company_revenue,
employee_size, region, funding_stage, strategic_fit, revenue_potential, regulatory_pressure,
cybersecurity_criticality, value_add, index_score, priority, sales_action, rationale
"""


def score_chunk(companies):
    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": build_prompt(companies)}],
        temperature=0.2,
        max_tokens=4000,
    )
    raw   = response.choices[0].message.content.strip()
    raw   = re.sub(r"```json|```", "", raw).strip()
    match = re.search(r"\[.*\]", raw, re.DOTALL)
    if match:
        raw = match.group(0)
    return json.loads(raw)


def score_all(df, prog, status, chunk_size=5):
    companies   = df.to_dict(orient="records")
    all_results = []
    total       = len(companies)

    for i in range(0, total, chunk_size):
        chunk = companies[i: i + chunk_size]
        end   = min(i + chunk_size, total)
        pct   = int((end / total) * 100)
        status.markdown(f"<p style='font-family:DM Mono,monospace;font-size:0.75rem;color:#5a5a5a;'>> PROCESSING {i+1}–{end} / {total}</p>", unsafe_allow_html=True)
        prog.progress(pct)
        try:
            results = score_chunk(chunk)
            all_results.extend(results)
        except Exception as e:
            st.warning(f"Batch {i+1}–{end} error: {e}")

    status.markdown("<p style='font-family:DM Mono,monospace;font-size:0.75rem;color:#FF4040;'>> ANALYSIS COMPLETE</p>", unsafe_allow_html=True)
    prog.progress(100)

    if not all_results:
        return pd.DataFrame()

    df_r = pd.DataFrame(all_results)
    if "index_score" in df_r.columns:
        df_r = df_r.sort_values("index_score", ascending=False).reset_index(drop=True)
    return df_r


def export_excel(df):
    output = BytesIO()

    output_columns = [
        "company_name", "website", "specialities", "industry",
        "company_revenue", "employee_size", "region", "funding_stage",
        "strategic_fit", "revenue_potential", "regulatory_pressure",
        "cybersecurity_criticality", "value_add", "index_score",
        "priority", "sales_action", "rationale",
    ]

    headers = [
        "Company Name", "Website", "Company Specialities", "Industry",
        "Company Revenue", "Employee Size", "Region", "Funding Stage",
        "Strategic Fit (0-10)", "Revenue Potential (0-10)",
        "Regulatory Pressure (0-10)", "Cybersecurity Criticality (0-10)",
        "Value-Add / Differentiation (0-10)", "Total Index Score (0-50)",
        "Priority for Outreach", "Sales Action", "Rationale",
    ]

    export_data = []
    for _, row in df.iterrows():
        export_data.append([str(row.get(col, "")) for col in output_columns])

    df_export = pd.DataFrame(export_data, columns=headers)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Scored Results")
        ws = writer.sheets["Scored Results"]

        thin = Border(
            left=Side(style='thin',   color='CCCCCC'),
            right=Side(style='thin',  color='CCCCCC'),
            top=Side(style='thin',    color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        WHITE     = PatternFill(fill_type="solid", fgColor="FFFFFF")
        LIGHTGREY = PatternFill(fill_type="solid", fgColor="F7F7F7")

        HEADER_FILL = PatternFill(fill_type="solid", fgColor="0D0D33")
        for cn in range(1, len(headers) + 1):
            c           = ws.cell(row=1, column=cn)
            c.fill      = HEADER_FILL
            c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = thin
        ws.row_dimensions[1].height = 40

        priority_colors = {
            "Very High": ("00BF59", "FFFFFF"),
            "High":      ("FFD900", "000000"),
            "Medium":    ("FF9900", "FFFFFF"),
            "Low":       ("E63333", "FFFFFF"),
        }

        def score_color(s):
            if s >= 40:   return ("00BF59", "FFFFFF")
            elif s >= 30: return ("FFD900", "000000")
            elif s >= 20: return ("FF9900", "FFFFFF")
            else:         return ("E63333", "FFFFFF")

        SCORE_COL    = 14
        PRIORITY_COL = 15
        TOTAL_COLS   = len(headers)

        for rn in range(2, len(df_export) + 2):
            for cn in range(1, TOTAL_COLS + 1):
                c           = ws.cell(row=rn, column=cn)
                c.fill      = WHITE if rn % 2 == 0 else LIGHTGREY
                c.font      = Font(bold=False, color="1A1A1A", size=10, name="Calibri")
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                c.border    = thin
            ws.row_dimensions[rn].height = 18

            pval = str(ws.cell(row=rn, column=PRIORITY_COL).value).strip()
            if pval in priority_colors:
                bg, fg      = priority_colors[pval]
                c           = ws.cell(row=rn, column=PRIORITY_COL)
                c.fill      = PatternFill(fill_type="solid", fgColor=bg)
                c.font      = Font(bold=True, color=fg, size=10, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")

            try:
                sval        = int(ws.cell(row=rn, column=SCORE_COL).value)
                bg, fg      = score_color(sval)
                c           = ws.cell(row=rn, column=SCORE_COL)
                c.fill      = PatternFill(fill_type="solid", fgColor=bg)
                c.font      = Font(bold=True, color=fg, size=11, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            except:
                pass

        col_widths = {
            1: 22, 2: 28, 3: 35, 4: 18, 5: 16,
            6: 14, 7: 16, 8: 18, 9: 12, 10: 14,
            11: 16, 12: 20, 13: 20, 14: 14,
            15: 16, 16: 30, 17: 45,
        }
        for cn, width in col_widths.items():
            ws.column_dimensions[get_column_letter(cn)].width = width

        ws.freeze_panes = "A2"

    output.seek(0)
    return output


def main():
    st.markdown("""
    <div class="topnav">
        <div class="topnav-left">
            <div class="topnav-brand">
                <span class="topnav-dot"></span>
                EBRYX
            </div>
            <div class="topnav-divider"></div>
            <div class="topnav-tag">Lead Intelligence Platform</div>
        </div>
        <div class="topnav-right">
            <div class="live-dot"></div>
            GROQ LLAMA 3.3 · 70B
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="hero">
        <div class="hero-bg-text">SCORE</div>
        <div class="hero-label">AI Sales Intelligence</div>
        <div class="hero-h1">
            SCORE &amp;<br>
            <span class="red">PRIORITIZE</span><br>
            YOUR LEADS
        </div>
        <p class="hero-p">
            Upload your prospect list. Our AI engine researches, scores, and ranks every
            company across 5 strategic dimensions — ready to download in seconds.
        </p>
    </div>
    """, unsafe_allow_html=True)

    col_l, col_r = st.columns([1, 2.4], gap="small")

    with col_l:
        st.markdown("<div style='padding:28px 4px 28px 0;'>", unsafe_allow_html=True)
        st.markdown('<div class="sec-hdr">Input</div>', unsafe_allow_html=True)

        st.markdown("""
        <div class="req-card">
            <div class="req-card-title">Required Columns</div>
            <span class="col-pill">Company Name</span>
            <span class="col-pill">Website</span>
            <div style="margin-top:10px;font-size:0.72rem;color:#3a3a3a;line-height:1.6;">
                Additional columns are passed to the AI for richer scoring context.
            </div>
        </div>
        """, unsafe_allow_html=True)

        uploaded   = st.file_uploader("UPLOAD FILE", type=["xlsx", "xls", "csv"], label_visibility="visible")
        df_input   = None
        run_button = False

        if uploaded:
            try:
                df_input = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
                df_input.columns = df_input.columns.str.strip()
                df_input = df_input.dropna(how="all").reset_index(drop=True)

                st.markdown(f"""
                <div style="background:rgba(196,26,26,0.06);border:1px solid rgba(196,26,26,0.2);
                            border-left:2px solid #C41A1A;border-radius:4px;padding:14px 16px;margin:14px 0;">
                    <div style="font-family:'DM Mono',monospace;font-size:0.6rem;color:#FF4040;
                                letter-spacing:0.15em;text-transform:uppercase;margin-bottom:6px;">✓ File Loaded</div>
                    <div style="font-family:'Bebas Neue',sans-serif;font-size:2.4rem;color:#FFFFFF;line-height:1;">{len(df_input)}</div>
                    <div style="font-family:'DM Mono',monospace;font-size:0.62rem;color:#5a5a5a;
                                letter-spacing:0.1em;text-transform:uppercase;">Companies Detected</div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown('<div class="sec-hdr" style="margin-top:18px;">Preview</div>', unsafe_allow_html=True)
                st.dataframe(df_input.head(4), use_container_width=True, height=150)

                st.markdown("<div style='margin-top:18px;'>", unsafe_allow_html=True)
                run_button = st.button("▶  RUN SCORING ENGINE", use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

            except Exception as e:
                st.error(f"Error: {e}")
        else:
            st.markdown("""
            <div class="up-placeholder">
                <div style="font-size:0.7rem;color:#3a3a3a;font-family:'DM Mono',monospace;
                            letter-spacing:0.1em;line-height:2.2;">
                    DRAG & DROP OR CLICK TO UPLOAD<br>
                    <span style="color:#2a2a2a">.XLSX · .XLS · .CSV</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("</div>", unsafe_allow_html=True)

    with col_r:
        st.markdown("<div style='padding:28px 0 28px 16px;'>", unsafe_allow_html=True)
        st.markdown('<div class="sec-hdr">Results</div>', unsafe_allow_html=True)

        if run_button and df_input is not None:
            prog   = st.progress(0)
            status = st.empty()
            df_res = score_all(df_input, prog, status)
            st.session_state["df_results"] = df_res

        if "df_results" not in st.session_state:
            st.markdown("""
            <div class="empty-st">
                <div class="empty-icon">[ _ ]</div>
                <div class="empty-title">NO DATA YET</div>
                <div class="empty-sub">Upload a file and run the scoring engine to see AI-powered lead intelligence.</div>
            </div>
            """, unsafe_allow_html=True)

        elif not st.session_state["df_results"].empty:
            df_r = st.session_state["df_results"]

            t  = len(df_r)
            vh = len(df_r[df_r["index_score"] >= 40])
            hi = len(df_r[(df_r["index_score"] >= 30) & (df_r["index_score"] < 40)])
            md = len(df_r[(df_r["index_score"] >= 20) & (df_r["index_score"] < 30)])
            lo = len(df_r[df_r["index_score"] < 20])

            st.markdown(f"""
            <div class="metrics-row">
                <div class="met-card c-total"><div class="met-num" style="color:#ffffff">{t}</div><div class="met-lbl">Total</div></div>
                <div class="met-card c-vh"><div class="met-num" style="color:#FF4040">{vh}</div><div class="met-lbl">Very High</div></div>
                <div class="met-card c-high"><div class="met-num" style="color:#FF8C00">{hi}</div><div class="met-lbl">High</div></div>
                <div class="met-card c-med"><div class="met-num" style="color:#A0A0A0">{md}</div><div class="met-lbl">Medium</div></div>
                <div class="met-card c-low"><div class="met-num" style="color:#3a3a3a">{lo}</div><div class="met-lbl">Low</div></div>
            </div>
            """, unsafe_allow_html=True)

            f1, f2 = st.columns([2, 1])
            with f1:
                pf = st.multiselect(
                    "FILTER BY PRIORITY",
                    ["Very High", "High", "Medium", "Low"],
                    default=["Very High", "High", "Medium", "Low"]
                )
            with f2:
                sb = st.selectbox(
                    "SORT BY",
                    ["index_score", "strategic_fit", "revenue_potential",
                     "regulatory_pressure", "cybersecurity_criticality", "value_add"],
                    format_func=lambda x: x.replace("_", " ").upper()
                )

            df_d  = df_r[df_r["priority"].isin(pf)].sort_values(sb, ascending=False).reset_index(drop=True)
            cols  = ["company_name", "industry", "region", "employee_size",
                     "strategic_fit", "revenue_potential", "regulatory_pressure",
                     "cybersecurity_criticality", "value_add", "index_score", "priority"]
            avail = [c for c in cols if c in df_d.columns]

            st.dataframe(
                df_d[avail].rename(columns={
                    "company_name":              "Company",
                    "industry":                  "Industry",
                    "region":                    "Region",
                    "employee_size":             "Headcount",
                    "strategic_fit":             "S.Fit",
                    "revenue_potential":         "Revenue",
                    "regulatory_pressure":       "Reg.",
                    "cybersecurity_criticality": "Cyber",
                    "value_add":                 "Value",
                    "index_score":               "Score",
                    "priority":                  "Priority",
                }),
                use_container_width=True,
                height=300
            )

            st.markdown("<div style='margin:16px 0;'>", unsafe_allow_html=True)
            excel_data = export_excel(df_r)
            st.download_button(
                label="↓  DOWNLOAD SCORED EXCEL",
                data=excel_data,
                file_name="ebryx_scored_leads.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown('<div class="sec-hdr" style="margin-top:8px;">Company Deep-Dive</div>', unsafe_allow_html=True)

            for _, row in df_d.iterrows():
                score    = row.get("index_score", 0)
                priority = row.get("priority", "N/A")
                company  = row.get("company_name", "N/A")

                with st.expander(f"  {company}   ·   {score}/50   ·   {priority}"):
                    d1, d2, d3, d4 = st.columns(4)
                    for col_obj, lbl, val in [
                        (d1, "Industry",  row.get("industry", "—")),
                        (d2, "Revenue",   row.get("company_revenue", "—")),
                        (d3, "Headcount", row.get("employee_size", "—")),
                        (d4, "Funding",   row.get("funding_stage", "—")),
                    ]:
                        with col_obj:
                            st.markdown(f'<div class="detail-label">{lbl}</div><div class="detail-value">{val}</div>', unsafe_allow_html=True)

                    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

                    bars = ""
                    for lbl, key in [
                        ("Strategic Fit",     "strategic_fit"),
                        ("Revenue Potential", "revenue_potential"),
                        ("Regulatory Press.", "regulatory_pressure"),
                        ("Cyber Criticality", "cybersecurity_criticality"),
                        ("Value-Add",         "value_add"),
                    ]:
                        v   = row.get(key, 0)
                        pct = (v / 10) * 100
                        c   = score_color_ui(v)
                        bars += f'<div class="sb-row"><span class="sb-lbl">{lbl}</span><div class="sb-track"><div class="sb-fill" style="width:{pct}%;background:{c};"></div></div><span class="sb-val" style="color:{c}">{v}</span></div>'

                    st.markdown(f'<div style="background:var(--black-3);border:1px solid var(--border);border-top:1px solid var(--border-red);border-radius:4px;padding:16px 20px;margin-bottom:12px;">{bars}</div>', unsafe_allow_html=True)

                    r1, r2 = st.columns(2)
                    with r1:
                        st.markdown(f'<div class="rationale-box"><b style="font-size:0.62rem;color:#3a3a3a;letter-spacing:0.15em;text-transform:uppercase;font-family:DM Mono,monospace;">Rationale</b><br><br>{row.get("rationale","—")}</div>', unsafe_allow_html=True)
                    with r2:
                        st.markdown(f'<div class="action-box"><b style="font-size:0.62rem;color:#C41A1A;letter-spacing:0.15em;text-transform:uppercase;font-family:DM Mono,monospace;">Sales Action</b><br><br>{row.get("sales_action","—")}</div>', unsafe_allow_html=True)

        st.markdown("</div>", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
